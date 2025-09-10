import os
import re
import json
import random
import zipfile
import tempfile
import shutil
import threading
import time
import datetime
from datetime import timedelta
from io import BytesIO

import requests
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.core import serializers
from django.core.cache import cache
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.files.temp import NamedTemporaryFile
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.mail import send_mail
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.contrib import messages
from django.contrib.auth import (
    authenticate, login as login_auth, logout, update_session_auth_hash
)
from django.contrib.auth.forms import AuthenticationForm, PasswordChangeForm
from django.contrib.auth.decorators import login_required

from django.db.models import (
    Prefetch, Q, Count, Avg, Sum, Max, Min, F,
    ExpressionWrapper, DurationField
)

from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST

from blog import models
from blog.models import yazi, category, SiteContent
from .models import Ogrenci, Ders, EzberSuresi, DersNotu, EzberKaydi, SinavSonucu
from .models import Alinti

# Global restore progress değişkeni
restore_progress = {
    'status': 'not_started',
    'progress': 0,
    'message': 'Geri yükleme başlatılmadı'
}

@login_required(login_url='login')
def backup_data(request):
    """
    Tüm verileri yedekler + fotoğrafları ayrı dizine kopyalar
    """
    try:
        # Yedekleme klasörünü oluştur
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups', f'backup_{timestamp}')
        os.makedirs(backup_dir, exist_ok=True)
        
        # Fotoğraflar için alt dizin oluştur
        photos_dir = os.path.join(backup_dir, 'photos')
        os.makedirs(photos_dir, exist_ok=True)
        
        # Resim bilgilerini depolamak için liste
        photo_info = []
        
        # Yazı resimlerini kopyala
        yazilar = yazi.objects.all()
        for yazi_obj in yazilar:
            if yazi_obj.imageUrl:
                try:
                    source_path = yazi_obj.imageUrl.path
                    if os.path.exists(source_path):
                        filename = os.path.basename(source_path)
                        dest_path = os.path.join(photos_dir, filename)
                        
                        # Dosyayı kopyala
                        shutil.copy2(source_path, dest_path)
                        
                        photo_info.append({
                            'type': 'yazi',
                            'id': yazi_obj.id,
                            'filename': filename,
                            'field': 'imageUrl'
                        })
                        print(f"Yazı {yazi_obj.id} resmi kopyalandı: {filename}")
                except Exception as e:
                    print(f"Yazı {yazi_obj.id} resim kopyalama hatası: {str(e)}")
        
        # Öğrenci profil fotoğraflarını kopyala
        ogrenciler = Ogrenci.objects.all()
        for ogrenci in ogrenciler:
            if ogrenci.profil_foto:
                try:
                    source_path = ogrenci.profil_foto.path
                    if os.path.exists(source_path):
                        filename = os.path.basename(source_path)
                        dest_path = os.path.join(photos_dir, filename)
                        
                        # Dosyayı kopyala
                        shutil.copy2(source_path, dest_path)
                        
                        photo_info.append({
                            'type': 'ogrenci',
                            'id': ogrenci.id,
                            'filename': filename,
                            'field': 'profil_foto'
                        })
                        print(f"Öğrenci {ogrenci.id} resmi kopyalandı: {filename}")
                except Exception as e:
                    print(f"Öğrenci {ogrenci.id} resim kopyalama hatası: {str(e)}")
        
        # Verileri JSON'a yaz (Alinti modelini de ekle)
        backup_data = {
            'ogrenciler': serializers.serialize('json', ogrenciler),
            'yazilar': serializers.serialize('json', yazilar),
            'ezber_kayitlari': serializers.serialize('json', EzberKaydi.objects.all()),
            'sinav_sonuclari': serializers.serialize('json', SinavSonucu.objects.all()),
            'ders_notlari': serializers.serialize('json', DersNotu.objects.all()),
            'alintilar': serializers.serialize('json', Alinti.objects.all()),  # Yeni eklenen
            'dersler': serializers.serialize('json', Ders.objects.all()),      # Dersler
            'ezber_sureleri': serializers.serialize('json', EzberSuresi.objects.all()),  # EzberSuresi
            'photo_info': photo_info,
            'backup_date': timezone.now().isoformat(),
            'backup_version': '1.3'  # Versiyonu güncelle
        }
        
        # JSON dosyasını kaydet
        json_path = os.path.join(backup_dir, 'backup.json')
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2)
        
        # ZIP dosyası oluştur
        zip_filename = f'backup_{timestamp}.zip'
        zip_path = os.path.join(settings.MEDIA_ROOT, 'backups', zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            # JSON dosyasını ekle
            zipf.write(json_path, 'backup.json')
            # Fotoğrafları ekle
            for root, dirs, files in os.walk(photos_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    zipf.write(file_path, os.path.join('photos', file))
        
        # Geçici klasörü temizle
        shutil.rmtree(backup_dir)
        
        # İndirme için hazırla
        with open(zip_path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
        
        messages.success(request, 'Veri yedeği başarıyla oluşturuldu ve indirildi.')
        return response
        
    except Exception as e:
        messages.error(request, f'Yedekleme sırasında hata oluştu: {str(e)}')
        return redirect('list_backups')

@login_required(login_url='login')
def list_backups(request):
    """
    Mevcut yedekleri listeler
    """
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    backups = []
    total_size = 0
    photo_count = 0
    
    if os.path.exists(backup_dir):
        for filename in os.listdir(backup_dir):
            if filename.endswith('.zip'):
                filepath = os.path.join(backup_dir, filename)
                file_time = os.path.getmtime(filepath)
                file_size = os.path.getsize(filepath)
                
                # ZIP içindeki fotoğraf sayısını say (yaklaşık)
                try:
                    with zipfile.ZipFile(filepath, 'r') as zipf:
                        photo_files = [f for f in zipf.namelist() if f.startswith('photos/') and not f.endswith('/')]
                        photo_count += len(photo_files)
                except:
                    photo_count = 0
                
                backups.append({
                    'filename': filename,
                    'filepath': filepath,
                    'date': timezone.datetime.fromtimestamp(file_time),
                    'size': file_size
                })
                total_size += file_size
    
    # Tarihe göre sırala (yeniden eskiye)
    backups.sort(key=lambda x: x['date'], reverse=True)
    
    return render(request, 'backup_list.html', {
        'backups': backups,
        'total_size': total_size,
        'photo_count': photo_count
    })

@login_required(login_url='login')
def restore_data(request):
    """
    Geri yükleme sayfasını gösterir veya işlemi başlatır
    """
    if request.method == 'GET':
        # GET isteği için yedek listesini göster
        backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
        backups = []
        total_size = 0
        photo_count = 0
        
        if os.path.exists(backup_dir):
            for filename in os.listdir(backup_dir):
                if filename.endswith('.zip'):
                    filepath = os.path.join(backup_dir, filename)
                    file_time = os.path.getmtime(filepath)
                    file_size = os.path.getsize(filepath)
                    
                    backups.append({
                        'filename': filename,
                        'filepath': filepath,
                        'date': timezone.datetime.fromtimestamp(file_time),
                        'size': file_size
                    })
                    total_size += file_size
        
        backups.sort(key=lambda x: x['date'], reverse=True)
        
        return render(request, 'restore_data.html', {
            'backups': backups,
            'total_size': total_size,
            'photo_count': photo_count
        })
    
    # POST isteği için geri yükleme işlemini başlat
    global restore_progress
    
    if 'backup_file' not in request.FILES:
        messages.error(request, 'Lütfen bir yedek dosyası seçin.')
        return redirect('restore_data')
    
    backup_file = request.FILES['backup_file']
    
    # İlerleme durumunu sıfırla
    restore_progress = {
        'status': 'started',
        'progress': 0,
        'message': 'Yedek dosyası işleniyor...'
    }
    
    try:
        # Geçici dosyayı kaydet
        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp_restore')
        os.makedirs(temp_dir, exist_ok=True)
        
        zip_path = os.path.join(temp_dir, 'backup.zip')
        with open(zip_path, 'wb+') as destination:
            for chunk in backup_file.chunks():
                destination.write(chunk)
        
        # Arka planda geri yükleme işlemini başlat
        thread = threading.Thread(target=restore_backup_process, args=(zip_path,))
        thread.daemon = True
        thread.start()
        
        messages.info(request, 'Geri yükleme işlemi başlatıldı. Lütfen bekleyin...')
        return redirect('restore_data')
        
    except Exception as e:
        restore_progress = {
            'status': 'error',
            'progress': 0,
            'message': f'Hata: {str(e)}'
        }
        messages.error(request, f'Geri yükleme başlatılamadı: {str(e)}')
        return redirect('restore_data')

@login_required(login_url='login')
def restore_progress_api(request):
    """Geri yükleme ilerleme durumunu JSON olarak döndürür"""
    global restore_progress
    return JsonResponse(restore_progress)

def restore_backup_process(zip_path):
    """Geri yükleme işlemini yürütür (arka plan thread'inde çalışır)"""
    global restore_progress
    
    try:
        # 1. Adım: Dosya doğrulama
        update_restore_progress(10, 'Yedek dosyası doğrulanıyor...')
        time.sleep(1)
        
        if not zipfile.is_zipfile(zip_path):
            raise ValueError('Geçerli bir ZIP dosyası değil')
        
        # 2. Adım: ZIP'i aç
        update_restore_progress(20, 'Yedek dosyası açılıyor...')
        time.sleep(1)
        
        extract_dir = tempfile.mkdtemp()
        
        try:
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)
            
            # 3. Adım: JSON dosyasını oku
            update_restore_progress(30, 'Yedek verileri okunuyor...')
            time.sleep(1)
            
            json_path = os.path.join(extract_dir, 'backup.json')
            if not os.path.exists(json_path):
                raise ValueError('Yedek dosyasında backup.json bulunamadı')
            
            with open(json_path, 'r', encoding='utf-8') as f:
                backup_data = json.load(f)
            
            # 4. Adım: Fotoğrafları yükle
            update_restore_progress(40, 'Fotoğraflar yükleniyor...')
            time.sleep(1)
            
            photo_info = backup_data.get('photo_info', [])
            photo_mappings = {}
            photos_dir = os.path.join(extract_dir, 'photos')
            
            if os.path.exists(photos_dir):
                for photo in photo_info:
                    photo_filename = photo['filename']
                    photo_path = os.path.join(photos_dir, photo_filename)
                    
                    if os.path.exists(photo_path):
                        with open(photo_path, 'rb') as f:
                            photo_content = f.read()
                        photo_mappings[(photo['type'], photo['id'])] = {
                            'content': photo_content,
                            'filename': photo_filename
                        }
            
            # 5. Adım: Mevcut verileri yedekle (önlem amaçlı)
            update_restore_progress(50, 'Mevcut veriler yedekleniyor...')
            time.sleep(2)
            
            create_emergency_backup()
            
            # 6. Adım: Veritabanı constraint'lerini devre dışı bırak
            update_restore_progress(55, 'Veritabanı constraint\'leri devre dışı bırakılıyor...')
            time.sleep(1)
            
            from django.db import connection
            with connection.cursor() as cursor:
                cursor.execute('PRAGMA foreign_keys=OFF;')
            
            # 7. Adım: Verileri geri yükle - Tüm modelleri DOĞRU SIRADA temizle
            update_restore_progress(60, 'Mevcut veriler temizleniyor...')
            time.sleep(2)
            
            # FOREIGN KEY ilişkilerine göre doğru sırada silme
            # Önce foreign key bağımlılığı olan modeller
            EzberKaydi.objects.all().delete()
            SinavSonucu.objects.all().delete()
            DersNotu.objects.all().delete()
            Alinti.objects.all().delete()
            EzberSuresi.objects.all().delete()
            
            # Sonra diğer modeller
            yazi.objects.all().delete()
            Ogrenci.objects.all().delete()
            Ders.objects.all().delete()
            
            # 8. Adım: Temel modelleri geri yükle (ters sırada)
            update_restore_progress(65, 'Temel veriler geri yükleniyor...')
            time.sleep(2)
            
            # Önce Dersler (diğer modeller buna bağlı olabilir)
            for obj in serializers.deserialize('json', backup_data['dersler']):
                obj.save()
            
            # 9. Adım: Öğrencileri geri yükle
            update_restore_progress(70, 'Öğrenci verileri geri yükleniyor...')
            time.sleep(2)
            
            for obj in serializers.deserialize('json', backup_data['ogrenciler']):
                ogrenci = obj.object
                # Eğer fotoğraf eşleşmesi varsa
                photo_key = ('ogrenci', ogrenci.id)
                if photo_key in photo_mappings:
                    try:
                        photo_data = photo_mappings[photo_key]
                        ogrenci.profil_foto.save(
                            photo_data['filename'], 
                            ContentFile(photo_data['content']), 
                            save=False
                        )
                    except Exception as e:
                        print(f"Öğrenci fotoğraf yükleme hatası: {str(e)}")
                ogrenci.save()
            
            # 10. Adım: Yazıları geri yükle
            update_restore_progress(75, 'Yazılar geri yükleniyor...')
            time.sleep(2)
            
            for obj in serializers.deserialize('json', backup_data['yazilar']):
                yazi_obj = obj.object
                # Eğer fotoğraf eşleşmesi varsa
                photo_key = ('yazi', yazi_obj.id)
                if photo_key in photo_mappings:
                    try:
                        photo_data = photo_mappings[photo_key]
                        yazi_obj.imageUrl.save(
                            photo_data['filename'], 
                            ContentFile(photo_data['content']), 
                            save=False
                        )
                    except Exception as e:
                        print(f"Yazı fotoğraf yükleme hatası: {str(e)}")
                yazi_obj.save()
            
            # 11. Adım: Diğer modelleri geri yükle
            update_restore_progress(80, 'Diğer veriler geri yükleniyor...')
            time.sleep(2)
            
            # EzberSuresi
            for obj in serializers.deserialize('json', backup_data['ezber_sureleri']):
                obj.save()
            
            # Alıntılar
            for obj in serializers.deserialize('json', backup_data['alintilar']):
                obj.save()
            
            # EzberKaydi
            for obj in serializers.deserialize('json', backup_data['ezber_kayitlari']):
                obj.save()
            
            # SınavSonuçları
            for obj in serializers.deserialize('json', backup_data['sinav_sonuclari']):
                obj.save()
            
            # DersNotları
            for obj in serializers.deserialize('json', backup_data['ders_notlari']):
                obj.save()
            
            # 12. Adım: Veritabanı constraint'lerini tekrar etkinleştir
            update_restore_progress(85, 'Veritabanı constraint\'leri etkinleştiriliyor...')
            time.sleep(1)
            
            with connection.cursor() as cursor:
                cursor.execute('PRAGMA foreign_keys=ON;')
                # Veritabanı bütünlüğünü kontrol et
                cursor.execute('PRAGMA integrity_check;')
                result = cursor.fetchone()
                if result and result[0] != 'ok':
                    raise ValueError(f'Veritabanı bütünlük hatası: {result[0]}')
            
            # 13. Adım: Temizlik
            update_restore_progress(90, 'Geçici dosyalar temizleniyor...')
            time.sleep(1)
            
            # Geçici dosyaları temizle
            shutil.rmtree(extract_dir, ignore_errors=True)
            os.unlink(zip_path)
            
            # 14. Adım: İşlem tamamlandı
            update_restore_progress(100, 'Geri yükleme başarıyla tamamlandı!')
            time.sleep(2)
            
        except Exception as e:
            # Hata durumunda constraint'leri tekrar etkinleştir
            try:
                with connection.cursor() as cursor:
                    cursor.execute('PRAGMA foreign_keys=ON;')
            except:
                pass
            
            # Hata durumunda temizlik
            try:
                shutil.rmtree(extract_dir, ignore_errors=True)
                os.unlink(zip_path)
            except:
                pass
            raise e
                
    except Exception as e:
        # Hata durumunda ilerlemeyi güncelle
        update_restore_progress(0, f'Hata: {str(e)}', 'error')
        
        # Hata durumunda emergency backup'tan geri yükle
        try:
            restore_from_emergency_backup()
        except Exception as restore_error:
            print(f"Emergency restore hatası: {restore_error}")

def update_restore_progress(progress, message, status='processing'):
    """İlerleme durumunu günceller"""
    global restore_progress
    restore_progress = {
        'status': status,
        'progress': progress,
        'message': message
    }
    time.sleep(0.5)

def create_emergency_backup():
    """Acil durum yedeği oluşturur"""
    try:
        emergency_dir = os.path.join(settings.MEDIA_ROOT, 'emergency_backup')
        os.makedirs(emergency_dir, exist_ok=True)
        
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        emergency_file = os.path.join(emergency_dir, f'emergency_{timestamp}.json')
        
        # Tüm modelleri içeren acil durum yedeği
        data = {
            'ogrenciler': serializers.serialize('json', Ogrenci.objects.all()),
            'yazilar': serializers.serialize('json', yazi.objects.all()),
            'ezber_kayitlari': serializers.serialize('json', EzberKaydi.objects.all()),
            'sinav_sonuclari': serializers.serialize('json', SinavSonucu.objects.all()),
            'ders_notlari': serializers.serialize('json', DersNotu.objects.all()),
            'alintilar': serializers.serialize('json', Alinti.objects.all()),
            'dersler': serializers.serialize('json', Ders.objects.all()),
            'ezber_sureleri': serializers.serialize('json', EzberSuresi.objects.all()),
            'backup_date': timezone.now().isoformat(),
        }
        
        with open(emergency_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            
    except Exception as e:
        print(f"Emergency backup hatası: {e}")

def restore_from_emergency_backup():
    """Acil durum yedeğinden geri yükler"""
    try:
        emergency_dir = os.path.join(settings.MEDIA_ROOT, 'emergency_backup')
        if not os.path.exists(emergency_dir):
            return
        
        # En son emergency backup'ı bul
        backup_files = [f for f in os.listdir(emergency_dir) if f.endswith('.json')]
        if not backup_files:
            return
        
        latest_backup = max(backup_files, key=lambda x: os.path.getctime(os.path.join(emergency_dir, x)))
        backup_path = os.path.join(emergency_dir, latest_backup)
        
        with open(backup_path, 'r', encoding='utf-8') as f:
            backup_data = json.load(f)
        
        # Mevcut verileri temizle (aynı sırayla)
        EzberKaydi.objects.all().delete()
        SinavSonucu.objects.all().delete()
        DersNotu.objects.all().delete()
        Alinti.objects.all().delete()
        yazi.objects.all().delete()
        Ogrenci.objects.all().delete()
        Ders.objects.all().delete()
        EzberSuresi.objects.all().delete()
        
        # Verileri geri yükle (aynı sırayla)
        for obj in serializers.deserialize('json', backup_data['dersler']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['ezber_sureleri']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['ogrenciler']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['yazilar']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['alintilar']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['ezber_kayitlari']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['sinav_sonuclari']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['ders_notlari']):
            obj.save()
            
    except Exception as e:
        print(f"Emergency restore hatası: {e}")

@login_required(login_url='login')
def download_backup(request, filename):
    """
    Belirli bir yedeği indir
    """
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    filepath = os.path.join(backup_dir, filename)
    
    if os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    
    messages.error(request, 'İstenen yedek dosyası bulunamadı.')
    return redirect('list_backups')

@login_required(login_url='login')
@require_POST
def delete_backup(request, filename):
    """
    Belirli bir yedeği sil
    """
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    filepath = os.path.join(backup_dir, filename)
    
    if os.path.exists(filepath):
        os.remove(filepath)
        messages.success(request, 'Yedek dosyası başarıyla silindi.')
    else:
        messages.error(request, 'İstenen yedek dosyası bulunamadı.')
    
    return redirect('list_backups')
    """Acil durum yedeği oluşturur"""
    try:
        emergency_dir = os.path.join(settings.MEDIA_ROOT, 'emergency_backup')
        os.makedirs(emergency_dir, exist_ok=True)
        
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        emergency_file = os.path.join(emergency_dir, f'emergency_{timestamp}.json')
        
        # Basit bir veritabanı yedeği
        data = {
            'ogrenciler': serializers.serialize('json', Ogrenci.objects.all()),
            'yazilar': serializers.serialize('json', yazi.objects.all()),
            'ezber_kayitlari': serializers.serialize('json', EzberKaydi.objects.all()),
            'sinav_sonuclari': serializers.serialize('json', SinavSonucu.objects.all()),
            'ders_notlari': serializers.serialize('json', DersNotu.objects.all()),
            'backup_date': timezone.now().isoformat(),
        }
        
        with open(emergency_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
            
    except Exception as e:
        print(f"Emergency backup hatası: {e}")

def restore_from_emergency_backup():
    """Acil durum yedeğinden geri yükler"""
    try:
        emergency_dir = os.path.join(settings.MEDIA_ROOT, 'emergency_backup')
        if not os.path.exists(emergency_dir):
            return
        
        # En son emergency backup'ı bul
        backup_files = [f for f in os.listdir(emergency_dir) if f.endswith('.json')]
        if not backup_files:
            return
        
        latest_backup = max(backup_files, key=lambda x: os.path.getctime(os.path.join(emergency_dir, x)))
        backup_path = os.path.join(emergency_dir, latest_backup)
        
        with open(backup_path, 'r', encoding='utf-8') as f:
            backup_data = json.load(f)
        
        # Mevcut verileri temizle
        Ogrenci.objects.all().delete()
        EzberKaydi.objects.all().delete()
        SinavSonucu.objects.all().delete()
        DersNotu.objects.all().delete()
        yazi.objects.all().delete()
        
        # Verileri geri yükle
        for obj in serializers.deserialize('json', backup_data['ogrenciler']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['yazilar']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['ezber_kayitlari']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['sinav_sonuclari']):
            obj.save()
        
        for obj in serializers.deserialize('json', backup_data['ders_notlari']):
            obj.save()
            
    except Exception as e:
        print(f"Emergency restore hatası: {e}")

@login_required(login_url='login')
def download_backup(request, filename):
    """
    Belirli bir yedeği indir
    """
    backup_dir = os.path.join(settings.MEDIA_ROOT, 'backups')
    filepath = os.path.join(backup_dir, filename)
    
    if os.path.exists(filepath):
        with open(filepath, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
    
    messages.error(request, 'İstenen yedek dosyası bulunamadı.')
    return redirect('list_backups')



def export_ogrenci_listesi_excel(request):
    # Tüm öğrencileri al, aynı filtreleri uygula
    tum_ogrenciler = Ogrenci.objects.all().order_by('-kayit_tarihi')
    
    arama_terimi = request.GET.get('q')
    if arama_terimi:
        tum_ogrenciler = tum_ogrenciler.filter(ad_soyad__icontains=arama_terimi)
    
    seviye_filtre = request.GET.get('seviye')
    if seviye_filtre:
        tum_ogrenciler = tum_ogrenciler.filter(seviye=seviye_filtre)
    
    # Workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Öğrenci Listesi"
    
    # Başlık satırı
    columns = ['Öğrenci Adı-Soyadı', 'Sınav Ortalaması', 'Tamamlanan Ezber', 'Toplam Ezber', 'Seviye', 'Kayıt Tarihi', 'Özel Notlar']
    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Veriler
    for ogrenci in tum_ogrenciler:
        row_num += 1
        # Her öğrenci için istatistikleri hesapla
        ortalama = SinavSonucu.objects.filter(ogrenci=ogrenci).aggregate(Avg('puan'))['puan__avg'] or 0
        tamamlanan_ezber = EzberKaydi.objects.filter(ogrenci=ogrenci, durum='TAMAMLANDI').count()
        
        row = [
            ogrenci.ad_soyad,
            ortalama,
            tamamlanan_ezber,
            13,  # Toplam ezber sayısı (sabit)
            ogrenci.get_seviye_display(),
            ogrenci.kayit_tarihi.strftime("%d.%m.%Y"),
            ogrenci.ozel_notlar or ""
        ]
        
        for col_num, cell_value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.alignment = Alignment(horizontal='center')
    
    # Sütun genişliklerini ayarla
    column_widths = [30, 20, 20, 15, 15, 15, 40]
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width
    
    # HttpResponse oluştur
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ogrenci_listesi_{}.xlsx'.format(datetime.datetime.now().strftime("%Y%m%d_%H%M"))
    
    wb.save(response)
    return response

def export_ogrenci_detay_excel(request, id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    
    # Workbook oluştur
    wb = Workbook()
    
    # 1. Öğrenci Bilgileri sayfası
    ws_info = wb.active
    ws_info.title = "Öğrenci Bilgileri"
    
    # Başlık
    ws_info.merge_cells('A1:B1')
    title_cell = ws_info['A1']
    title_cell.value = f"{ogrenci.ad_soyad} - Öğrenci Detay Raporu"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Öğrenci bilgilerini yaz
    ogrenci_bilgileri = [
        ['Ad-Soyad', ogrenci.ad_soyad],
        ['Seviye', ogrenci.get_seviye_display()],
        ['Kayıt Tarihi', ogrenci.kayit_tarihi.strftime("%d.%m.%Y")],
        ['Kursta Geçen Süre', f"{(timezone.now().date() - ogrenci.kayit_tarihi).days} gün"],
        ['Özel Notlar', ogrenci.ozel_notlar or ""]
    ]
    
    row_num = 3
    for bilgi in ogrenci_bilgileri:
        ws_info.cell(row=row_num, column=1, value=bilgi[0])
        ws_info.cell(row=row_num, column=1).font = Font(bold=True)
        ws_info.cell(row=row_num, column=2, value=bilgi[1])
        row_num += 1
    
    # 2. Sınav Sonuçları sayfası
    ws_sinav = wb.create_sheet(title="Sınav Sonuçları")
    
    # Başlık
    ws_sinav.merge_cells('A1:F1')
    title_cell = ws_sinav['A1']
    title_cell.value = "Sınav Sonuçları"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Sütun başlıkları
    columns = ['Ders', 'Sınav Tipi', 'Puan', 'Tarih', 'Açıklama', 'Durum']
    row_num = 3
    for col_num, column_title in enumerate(columns, 1):
        cell = ws_sinav.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Sınav verileri
    sinav_sonuclari = SinavSonucu.objects.filter(ogrenci=ogrenci).select_related('ders')
    row_num = 4
    for sinav in sinav_sonuclari:
        durum = ""
        if sinav.puan >= 85:
            durum = "Çok İyi"
        elif sinav.puan >= 70:
            durum = "İyi"
        elif sinav.puan >= 50:
            durum = "Orta"
        else:
            durum = "Zayıf"
            
        row = [
            sinav.ders.get_tur_display(),
            sinav.get_sinav_tipi_display(),
            sinav.puan,
            sinav.tarih.strftime("%d.%m.%Y") if sinav.tarih else '',
            sinav.aciklama or '',
            durum
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws_sinav.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal='center')
        row_num += 1
    
    # Sınav istatistikleri
    sinav_ortalamasi = sinav_sonuclari.aggregate(ortalama=Avg('puan'))['ortalama'] or 0
    en_yuksek_puan = sinav_sonuclari.aggregate(en_yuksek=Max('puan'))['en_yuksek'] or 0
    en_dusuk_puan = sinav_sonuclari.aggregate(en_dusuk=Min('puan'))['en_dusuk'] or 0
    
    ws_sinav.cell(row=row_num+2, column=1, value="İstatistikler").font = Font(bold=True)
    ws_sinav.cell(row=row_num+3, column=1, value="Ortalama Puan")
    ws_sinav.cell(row=row_num+3, column=2, value=sinav_ortalamasi)
    ws_sinav.cell(row=row_num+4, column=1, value="En Yüksek Puan")
    ws_sinav.cell(row=row_num+4, column=2, value=en_yuksek_puan)
    ws_sinav.cell(row=row_num+5, column=1, value="En Düşük Puan")
    ws_sinav.cell(row=row_num+5, column=2, value=en_dusuk_puan)
    
    # 3. Ezber Kayıtları sayfası
    ws_ezber = wb.create_sheet(title="Ezber Kayıtları")
    
    # Başlık
    ws_ezber.merge_cells('A1:G1')
    title_cell = ws_ezber['A1']
    title_cell.value = "Ezber Kayıtları"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Sütun başlıkları
    columns = ['Sıra', 'Sure', 'Durum', 'Başlama Tarihi', 'Bitiş Tarihi', 'Süre (Gün)', 'Yorum']
    row_num = 3
    for col_num, column_title in enumerate(columns, 1):
        cell = ws_ezber.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Ezber verileri
    ezber_kayitlari = EzberKaydi.objects.filter(ogrenci=ogrenci).select_related('sure').order_by('sure__sira')
    row_num = 4
    for ezber in ezber_kayitlari:
        # Ezber süresini hesapla
        sure_gun = 0
        if ezber.baslama_tarihi and ezber.bitis_tarihi:
            sure_gun = (ezber.bitis_tarihi - ezber.baslama_tarihi).days
            
        row = [
            ezber.sure.sira,
            ezber.sure.ad,
            ezber.get_durum_display(),
            ezber.baslama_tarihi.strftime("%d.%m.%Y") if ezber.baslama_tarihi else '',
            ezber.bitis_tarihi.strftime("%d.%m.%Y") if ezber.bitis_tarihi else '',
            sure_gun,
            ezber.yorum or ''
        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = ws_ezber.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal='center')
        row_num += 1
    
    # Ezber istatistikleri
    tamamlanan_ezberler = ezber_kayitlari.filter(durum='TAMAMLANDI').count()
    devam_eden_ezberler = ezber_kayitlari.filter(durum='DEVAM').count()
    toplam_ezber = EzberSuresi.objects.count()
    
    # Ortalama ezber süresi (tamamlananlar için)
    tamamlanan_ezber_kayitlari = ezber_kayitlari.filter(durum='TAMAMLANDI')
    ortalama_ezber_suresi = 0
    if tamamlanan_ezber_kayitlari.exists():
        toplam_gun = 0
        for ezber in tamamlanan_ezber_kayitlari:
            if ezber.baslama_tarihi and ezber.bitis_tarihi:
                toplam_gun += (ezber.bitis_tarihi - ezber.baslama_tarihi).days
        ortalama_ezber_suresi = toplam_gun / tamamlanan_ezber_kayitlari.count()
    
    ws_ezber.cell(row=row_num+2, column=1, value="İstatistikler").font = Font(bold=True)
    ws_ezber.cell(row=row_num+3, column=1, value="Tamamlanan Ezber")
    ws_ezber.cell(row=row_num+3, column=2, value=tamamlanan_ezberler)
    ws_ezber.cell(row=row_num+4, column=1, value="Devam Eden Ezber")
    ws_ezber.cell(row=row_num+4, column=2, value=devam_eden_ezberler)
    ws_ezber.cell(row=row_num+5, column=1, value="Toplam Ezber")
    ws_ezber.cell(row=row_num+5, column=2, value=toplam_ezber)
    ws_ezber.cell(row=row_num+6, column=1, value="Tamamlama Oranı")
    ws_ezber.cell(row=row_num+6, column=2, value=f"{(tamamlanan_ezberler/toplam_ezber*100):.1f}%" if toplam_ezber > 0 else "0%")
    ws_ezber.cell(row=row_num+7, column=1, value="Ortalama Ezber Süresi (Gün)")
    ws_ezber.cell(row=row_num+7, column=2, value=f"{ortalama_ezber_suresi:.1f}")
    
    # 4. Performans Analizi sayfası
    ws_analiz = wb.create_sheet(title="Performans Analizi")
    
    # Başlık
    ws_analiz.merge_cells('A1:B1')
    title_cell = ws_analiz['A1']
    title_cell.value = "Performans Analizi"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Sınıf karşılaştırması
    sinif_ortalamasi = SinavSonucu.objects.aggregate(ortalama=Avg('puan'))['ortalama'] or 0
    sinif_ezber_ortalamasi = EzberKaydi.objects.filter(durum='TAMAMLANDI').values('ogrenci').annotate(
        tamamlanan=Count('id')
    ).aggregate(ortalama=Avg('tamamlanan'))['ortalama'] or 0
    
    analiz_verileri = [
        ['Öğrenci Ortalaması', sinav_ortalamasi],
        ['Sınıf Ortalaması', sinif_ortalamasi],
        ['Fark', sinav_ortalamasi - sinif_ortalamasi],
        ['', ''],
        ['Tamamlanan Ezber', tamamlanan_ezberler],
        ['Sınıf Ortalaması (Ezber)', sinif_ezber_ortalamasi],
        ['Fark', tamamlanan_ezberler - sinif_ezber_ortalamasi],
        ['', ''],
        ['Önerilen Çalışma Süresi', 
         '8 saat/gün' if sinav_ortalamasi < 50 else
         '6 saat/gün' if sinav_ortalamasi < 60 else
         '4 saat/gün' if sinav_ortalamasi < 70 else
         '2 saat/gün' if sinav_ortalamasi < 80 else
         '1 saat/gün'],
        ['Hafızlık Potansiyeli', 
         'Yüksek' if tamamlanan_ezberler >= 10 else
         'Orta' if tamamlanan_ezberler >= 7 else
         'Düşük' if tamamlanan_ezberler >= 4 else
         'Belirsiz']
    ]
    
    row_num = 3
    for veri in analiz_verileri:
        ws_analiz.cell(row=row_num, column=1, value=veri[0])
        if veri[0]:  # Başlık satırları için
            ws_analiz.cell(row=row_num, column=1).font = Font(bold=True)
        ws_analiz.cell(row=row_num, column=2, value=veri[1])
        row_num += 1
    
    # Sütun genişliklerini manuel olarak ayarla (MergedCell hatasını önlemek için)
    column_widths = {
        'Öğrenci Bilgileri': {'A': 20, 'B': 30},
        'Sınav Sonuçları': {'A': 15, 'B': 15, 'C': 10, 'D': 12, 'E': 20, 'F': 15},
        'Ezber Kayıtları': {'A': 8, 'B': 20, 'C': 15, 'D': 15, 'E': 15, 'F': 12, 'G': 30},
        'Performans Analizi': {'A': 25, 'B': 20}
    }
    
    for sheet_name, widths in column_widths.items():
        if sheet_name == 'Öğrenci Bilgileri':
            ws = ws_info
        elif sheet_name == 'Sınav Sonuçları':
            ws = ws_sinav
        elif sheet_name == 'Ezber Kayıtları':
            ws = ws_ezber
        elif sheet_name == 'Performans Analizi':
            ws = ws_analiz
        
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width
    
    # Dosyayı kaydet
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{ogrenci.ad_soyad}_detay_raporu_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    wb.save(response)
    return response
def format_gemini_response(text):
    """
    Gemini API'den gelen metni düzgün HTML formatına dönüştürür
    """
    if not text:
        return text
    
    # HTML etiketlerini temizle (güvenlik için)
    import html
    text = html.escape(text)
    
    # Madde işaretlerini tespit et ve düzenle
    text = re.sub(r'\*\*(\d+\.\s+[^*]+)\*\*', r'<br><strong>\1</strong><br>', text)
    text = re.sub(r'\*\*([^*]+)\*\*', r'<strong>\1</strong>', text)
    
    # Yıldız işaretli maddeleri liste öğelerine dönüştür
    text = re.sub(r'\* (\d+\.\s+[^*]+)', r'<li>\1</li>', text)
    text = re.sub(r'\* ([^*]+)', r'<li>\1</li>', text)
    
    # Liste başlangıçlarını tespit et
    text = re.sub(r'(<li>.*?</li>(?:\s*<li>.*?</li>)+)', r'<ul>\1</ul>', text, flags=re.DOTALL)
    
    # Satır sonlarını <br> ile değiştir
    text = text.replace('\n', '<br>')
    text = re.sub(r'(<br>){3,}', '<br><br>', text)  # Fazla boşlukları temizle
    
    return mark_safe(text)

@login_required(login_url='login')
@csrf_exempt
def arama_motoru(request):
    if request.method == 'POST':
        # JSON verisi mi form verisi mi kontrol et
        if request.content_type == 'application/json':
            data = json.loads(request.body)
            sorgu = data.get('sorgu', '')
        else:
            sorgu = request.POST.get('sorgu', '')
        
        if not sorgu or len(sorgu.strip()) == 0:
            return JsonResponse({'error': 'Sorgu boş olamaz'}, status=400)
        
        # Önbellek anahtarı oluştur
        cache_key = f"gemini_{hash(sorgu)}"
        cached_response = cache.get(cache_key)
        
        if cached_response:
            return JsonResponse({
                'cevap': cached_response,
                'sorgu': sorgu,
                'cached': True
            })
        
        # Gemini API isteği
        url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"
        
        headers = {
            'Content-Type': 'application/json',
            'X-goog-api-key': settings.GEMINI_API_KEY
        }
        
        # Daha iyi formatlanmış yanıt almak için prompt'u optimize et
# Daha iyi formatlanmış yanıt almak için prompt'u optimize et
        prompt = (
            f"Şeyma adında birine cevap verir gibi yanıtla. Karşındaki kişi bir kuran öğretmeni ve hafız. İsmi Şeyma çok zeki, çok güzel, çok değerli"
            f"Seninle konuşan ve konuştuğun karşındaki kişi olan Şeyma, Amine Hatun Kuran Kursunda Hafızlık Hazırlık Öğretmeni"
            f"Kullanıcının sorusu: {sorgu}. "
            f"Cevabın samimi, dostane ve bilgilendirici olsun. "
            f"Lütfen yanıtını aşağıdaki kurallara göre formatla:\n"
            f"1. Başlıklar için **kalın** kullan\n"
            f"2. Maddeler için * işareti kullan\n"
            f"3. Her maddeyi yeni satırda başlat\n"
            f"4. Paragraflar arasında boşluk bırak\n"
            f"6. HTML etiketi kullanma, sadece * ve ** işaretleri kullan."
        )

        payload = {
            "contents": [
                {
                    "parts": [
                        {
                            "text": prompt  # Buradaki "prompt" artık tek bir metin dizesi olacak
                        }
                    ]
                }
            ],
            "generationConfig": {
                "temperature": 0.7,
                "topK": 40,
                "topP": 0.95,
                "maxOutputTokens": 1024,
            }
        }
        
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=30)
            response.raise_for_status()
            
            result = response.json()
            
            # Yanıtı çıkar
            if (result.get('candidates') and 
                len(result['candidates']) > 0 and 
                result['candidates'][0].get('content') and
                result['candidates'][0]['content'].get('parts') and
                len(result['candidates'][0]['content']['parts']) > 0):
                
                cevap = result['candidates'][0]['content']['parts'][0]['text']
                
                # Metni formatla
                formatted_cevap = format_gemini_response(cevap)
                
                # Önbelleğe al (1 saat)
                cache.set(cache_key, formatted_cevap, 3600)
                
                return JsonResponse({
                    'cevap': formatted_cevap,
                    'sorgu': sorgu,
                    'success': True
                })
            else:
                return JsonResponse({
                    'error': 'API yanıt formatı beklenen şekilde değil'
                }, status=500)
                
        except requests.exceptions.HTTPError as e:
            error_msg = f"HTTP hatası: {str(e)}"
            if hasattr(e, 'response') and e.response.status_code == 429:
                error_msg = "Şu anda çok fazla istek yapıldı. Lütfen bir süre sonra tekrar deneyin."
            return JsonResponse({'error': error_msg}, status=500)
            
        except requests.exceptions.Timeout:
            return JsonResponse({'error': 'İstek zaman aşımına uğradı. Lütfen tekrar deneyin.'}, status=408)
            
        except requests.exceptions.ConnectionError:
            return JsonResponse({'error': 'İnternet bağlantı hatası. Lütfen bağlantınızı kontrol edin.'}, status=503)
            
        except Exception as e:
            return JsonResponse({'error': f'Beklenmeyen bir hata oluştu: {str(e)}'}, status=500)
    
    # GET isteği için HTML sayfasını göster
    return render(request, 'arama_motoru.html')

def home(request):
    son_yazilar = yazi.objects.filter(isActive=True).order_by('-id')[:3]  # en son 3 aktif yazı
    yazilar = yazi.objects.filter(isActive=True)
    anasayfa_alt_metin = SiteContent.objects.filter(slug='anasayfa-alt-metin').first()

    number = len(yazilar)
    sozler = [
    "Gönül ne kahve ister ne kahvehane, gönül sohbet ister kahve bahane.",
    "Kendini bilen, Rabbini bilir.",
    "Ne olursan ol yine gel.",
    "Her şey üstüne gelip seni dayanamayacağın bir noktaya getirdiğinde, sakın vazgeçme.",
    "Kalp deniz, dil kıyıdır. Denizde ne varsa kıyıya o vurur.",
    "Susmak, bazen en güçlü çığlıktır.",
    "Unutma, karanlık olmasaydı yıldızları göremezdik.",
    "Yürümeye devam eden yol alır, düşse bile kalkar.",
    "Küçük adımlar büyük yollar açar.",
    "Dünya seni yıkmadan önce sen kendini inşa et.",
    "Bazen kaybetmek, en büyük kazançtır.",
    "Gerçek güç, vazgeçmemekte saklıdır.",
    "Her gün yeni bir başlangıçtır.",
    "Karanlıktan korkma, yıldızlar orada doğar.",
    "Zihin neye inanırsa beden ona ulaşır.",
    "Bir şey değişir, her şey değişir.",
    "Düşüncelerini değiştir, hayatın değişsin.",
    "Engeller, seni durdurmak için değil, yön vermek için vardır.",
    "Başarı, tekrar tekrar denemekten geçer.",
    "Hayat seni yıkarsa, yeniden inşa et.",
    "Cesaret, korkmamak değil; korkuya rağmen yürümektir.",
    "Bugünün acısı, yarının gücüdür.",
    "Asıl savaş, insanın kendi içindedir.",
    "Kendine inandığın gün, dünya da sana inanır.",
    "Bir umut yeter, karanlığı aydınlatmaya.",
    "Hayallerini küçümseyenlerden uzak dur.",
    "Her sabah yeni bir mucizedir.",
    "Diken olmadan gül olmaz.",
    "İmkânsız, sadece daha fazla çaba gerektirir.",
    "Bazen yavaş gitmek, doğru gitmektir.",
    "Yalnızlık, bazen en iyi öğretmendir.",
    "Zor zamanlar, güçlü insanlar yaratır.",
    "Hiçbir rüzgar, yönünü bilmeyene yardım edemez.",
    "Hayat, cesur olanları ödüllendirir.",
    "Bir gün değil, her gün başla.",
    "Yüzleşmeden geçmeyen sınav, öğrenilmez.",
    "Yol senin, yürümek de.",
    "Umutsuzluk yok, sadece dinlenme molası var.",
    "Yenilmek değil, vazgeçmek kaybettirir.",
    "Kelimeler köprü kurar, sessizlik duvar.",
    "Kendinle barış, her şeyle barış getirir.",
    "Başarı, konfor alanının dışında başlar.",
    "Bir ışık ol, karanlıkta kalanlara yol göster.",
    "Kalbinin götürdüğü yere git.",
    "Beklemek değil, harekete geçmek değiştirir.",
    "Hayat kısa, hayalin peşinden git.",
    "Kırıldığın yerden güçlenirsin.",
    "Her şeyin başı niyet.",
    "Gözlerinle değil, kalbinle gör.",
    "İçindeki sesi dinle, o hiç yalan söylemez.",
    "Başlamak için mükemmel olmak zorunda değilsin.",
    "Hayal etmek başarmanın yarısıdır.",
    "Başarı, tekrar tekrar denemekten vazgeçmemektir.",
    "Yol ne kadar zor olursa olsun, vazgeçmek çözüm değildir.",
    "Bir adım at, yol seni takip edecektir.",
    "Zirveye giden yol, cesaretle başlar.",
    "Yapabileceğine inan, zaten yarısını başarmışsındır.",
    "Küçük adımlar, büyük zaferlerin başlangıcıdır.",
    "Bugün attığın adım, yarının başarısını belirler.",
    "Yorulmak, pes etmek için değil; dinlenip devam etmek içindir.",
    "Engeller, kararlılıkla aşılmak içindir.",
    "Bir fikrin varsa, bir yolun da vardır.",
    "İmkânsız, sadece daha fazla çaba gerektirir.",
    "Karanlık günler geçer, ışığı bekle.",
    "Gerçek güç, pes etmediğin an ortaya çıkar.",
    "Başarı, cesur olanların ödülüdür.",
    "Denemediğin sürece kaybetmiş sayılmazsın.",
    "Her gün yeni bir başlangıçtır.",
    "Hayat, cesur olanları ödüllendirir.",
    "Risk almadan kazanç olmaz.",
    "Zorluklar, seni güçlü kılmak için vardır.",
    "Kendine inan, çünkü başka kimse senin yerine yaşayamaz.",
    "Ne kadar yavaş ilerlediğin önemli değil, durmadığın sürece başarırsın.",
    "Her şey seninle başlar.",
    "Hayat bir aynadır, gülümsersen gülümser.",
    "Değişim, seninle başlar.",
    "En karanlık an, şafağa en yakın andır.",
    "Hayat bir mücadeledir, sen de bir savaşçısın.",
    "Düşersen kalk, çünkü ilerlemek için yürümek gerekir.",
    "Her kayıp bir ders, her ders bir adımdır.",
    "Kazanmak istemek yetmez, harekete geçmek gerekir.",
    "Bugün yapamadığın şey, yarının hedefi olsun.",
    "Kendini küçümseme, içinde evrenler var.",
    "İnandığın yolda yürü, sonunda ödül seni bulur.",
    "Özgüven, en güçlü silahtır.",
    "Unutma, en büyük başarılar en derin yaralardan doğar.",
    "Umut, en karanlık anların ilacıdır.",
    "Başarının sırrı, disiplin ve sabırdır.",
    "Bir amacı olan insanın gücüne sınır koyulamaz.",
    "Sen değişirsen dünya değişir.",
    "Küçük başarıları kutla, büyük hedeflere hazırlan.",
    "Hiçbir şey yapmamaktan iyidir denemek.",
    "Kendini geliştir, çünkü zaman seni beklemez.",
    "İnançsız bir adım bile ilerleme getirir.",
    "Kimi zaman düşmek, doğru yolu bulmak için gereklidir.",
    "Her gün bir önceki seninle yarış.",
    "Hayatta kalmak değil, yaşamak hedefin olsun.",
    "Gerçek başarı, iç huzurla gelir.",
    "Sen yeter ki başla, gerisi gelir.",
    "Zihnin sınır tanımaz, yeter ki onu serbest bırak.",
    "Bugün bir şey yap, yarın teşekkür edeceksin."
    "Kim demiş gül yaşar dikenin himayesinde? Dikenin itibarı ancak gül sayesinde",
    "Sessizlik cevapları verir",
    "İnsan kalbiyle insandır, kalpsiz beden cesettir.",
    "Susmak bazen en gür sestir."
    ]
    rastgele_soz = random.choice(sozler)
    return render(request, 'index.html', {
        'son_yazilar': son_yazilar,
        'rastgele_soz': rastgele_soz,
        'anasayfa_alt_metin': anasayfa_alt_metin,
        'number':number
    })
def about(request):
    hakkimda = SiteContent.objects.filter(slug='hakkimda').first()
    return render(request,'hakkimda.html', {
        "hakkimda":hakkimda,
    })

def iletisim(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        message = request.POST.get('message')
        
        # E-posta içeriğini daha düzenli hale getirme
        subject = f"SEYMAA.COM -  Yeni İletişim Formu: {name}"
        email_message = f"""
        Ad Soyad: {name}
        E-posta: {email}
        
        Mesaj:
        {message}
        
        Bu mesaj {settings.SITE_NAME} iletişim formundan gönderilmiştir.
        """
        
        try:
            # Mail gönderiminden önce bilgileri kontrol et
            
            send_mail(
                subject=subject,
                message=email_message,
                from_email=settings.DEFAULROM_EMAIL,
                recipient_list=[settings.CONTACT_EMAIL],
                fail_silently=False,
            )
            messages.success(request, 'Mesajınız başarıyla gönderildi! En kısa sürede sizinle iletişime geçeceğim.')
            return redirect('iletisim')
            
        except Exception as e:
            error_msg = f"Mail gönderilemedi. Hata: {str(e)}"
            print(error_msg)  # Konsola hata detayını yaz
            messages.error(request, 'Mesajınız gönderilirken bir hata oluştu. Lütfen daha sonra tekrar deneyin.')
    
    return render(request, 'iletisim.html')


@login_required(login_url='login')
def admin_yazi_listesi(request):
    yazilar = yazi.objects.all()
    aktif_yazi_sayisi = yazi.objects.filter(isActive=True).count()
    return render(request, 'list.html', {'yazilar': yazilar,'aktif_yazi_sayisi':aktif_yazi_sayisi})

@login_required(login_url='login')
def yazi_guncelle(request, id):
    yazim = get_object_or_404(yazi, id=id)

    if request.method == 'POST':
        yazim.title = request.POST.get('baslik')
        yazim.description = request.POST.get('description')
        
        # DÜZELTME: Select elementinden gelen değeri doğru işleme
        isActive = request.POST.get('aktif')
        yazim.isActive = (isActive == "True")  # "True" string'i ile karşılaştır
        
        # Resim güncelleme
        if 'image' in request.FILES:
            yazim.imageUrl = request.FILES['image']
            
        yazim.save()
        return redirect('admin-yazi-listesi')

    return render(request, 'list_duzenle.html', {'yazi': yazim})

@login_required(login_url='login')
def admin_yazi_sil(request, id):
    yazi_obj = get_object_or_404(yazi, id=id)
    if request.method == 'POST':
        yazi_obj.delete()
        return redirect('admin-yazi-listesi')
    return render(request, 'yazi_sil_onay.html', {'yazi': yazi_obj})

@login_required(login_url='login')
def alinti_yaz(request):
    if request.method == 'POST':
        quote_text = request.POST.get('quote_text')
        author = request.POST.get('author')
        source = request.POST.get('source')
        category = request.POST.get('category')
        isActive = request.POST.get('isActive') == 'on'

        if not quote_text:
            messages.error(request, 'Alıntı metni boş olamaz.')
            return render(request, 'alinti_yaz.html')

        try:
            Alinti.objects.create(
                quote_text=quote_text,
                author=author,
                source=source,
                category=category,
                isActive=isActive
            )
            messages.success(request, 'Alıntı başarıyla eklendi.')
            return redirect('alinti-listesi')
        except Exception as e:
            messages.error(request, f'Alıntı eklenirken bir hata oluştu: {str(e)}')

    return render(request, 'alinti_yaz.html')

@login_required(login_url='login')
def alinti_listesi(request):
    alinti_list = Alinti.objects.all().order_by('-created_at')
    
    # Durum filtreleme
    durum = request.GET.get('durum')
    if durum == 'aktif':
        alinti_list = alinti_list.filter(isActive=True)
    elif durum == 'pasif':
        alinti_list = alinti_list.filter(isActive=False)
    
    # Kategori filtreleme
    kategori = request.GET.get('kategori')
    if kategori:
        alinti_list = alinti_list.filter(category=kategori)
    
    # Sayfalama
    sayfa = request.GET.get('sayfa', 1)
    paginator = Paginator(alinti_list, 15)  # Sayfa başına 15 alıntı
    
    try:
        alintilar = paginator.page(sayfa)
    except PageNotAnInteger:
        alintilar = paginator.page(1)
    except EmptyPage:
        alintilar = paginator.page(paginator.num_pages)
        
    return render(request, 'alinti_yonetim.html', {'alintilar': alintilar})

@login_required(login_url='login')
def alinti_duzenle(request, id):
    alinti = get_object_or_404(Alinti, id=id)
    
    if request.method == 'POST':
        # Form verilerini al
        quote_text = request.POST.get('quote_text')
        author = request.POST.get('author')
        source = request.POST.get('source')
        category = request.POST.get('category')
        is_active = request.POST.get('isActive') == 'on'  # Checkbox kontrolü
        
        # Validasyon
        if not quote_text:
            messages.error(request, 'Alıntı metni boş olamaz!')
            return render(request, 'alinti_duzenle.html', {'alinti': alinti})
        
        # Değerleri güncelle
        alinti.quote_text = quote_text
        alinti.author = author
        alinti.source = source
        alinti.category = category
        alinti.isActive = is_active
        
        try:
            alinti.save()
            messages.success(request, 'Alıntı başarıyla güncellendi.')
            return redirect('alinti-listesi')
        except Exception as e:
            messages.error(request, f'Alıntı güncellenirken bir hata oluştu: {str(e)}')
    
    return render(request, 'alinti_duzenle.html', {'alinti': alinti})
# Herkese açık alıntı listesi (sadece aktif olanlar)
def tum_alintilar(request):
    alinti_list = Alinti.objects.filter(isActive=True).order_by('-created_at')
    
    # Kategori filtreleme
    kategori = request.GET.get('kategori')
    if kategori and kategori != 'tum':
        alinti_list = alinti_list.filter(category=kategori)
    
    # Sayfalama
    sayfa = request.GET.get('sayfa', 1)
    paginator = Paginator(alinti_list, 10)
    
    try:
        alintilar = paginator.page(sayfa)
    except PageNotAnInteger:
        alintilar = paginator.page(1)
    except EmptyPage:
        alintilar = paginator.page(paginator.num_pages)
        
    return render(request, 'alintilar.html', {
        'alintilar': alintilar,
        'is_public': True  # Template'de kullanmak için
    })

@login_required(login_url='login')
def alinti_sil(request, id):
    alinti = get_object_or_404(Alinti, id=id)
    
    if request.method == 'POST':
        try:
            alinti.delete()
            messages.success(request, 'Alıntı başarıyla silindi.')
        except Exception as e:
            messages.error(request, f'Alıntı silinirken bir hata oluştu: {str(e)}')
        
        return redirect('alinti-listesi')
    
    return render(request, 'alinti_sil_onay.html', {'alinti': alinti})


from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Avg, Q
from blog.models import yazi
from mainproject.models import Ogrenci, EzberKaydi, SinavSonucu

@login_required(login_url='login')
def admin_dashboard(request):
    # İstatistik verileri
    toplam_yazi = yazi.objects.count()
    toplam_ogrenci = Ogrenci.objects.count()
    
    # Sınıf ortalaması
    sinif_ortalamasi = SinavSonucu.objects.aggregate(
        ortalama=Avg('puan')
    )['ortalama'] or 0
    
    # Ezber istatistikleri
    tamamlanan_ezber = EzberKaydi.objects.filter(durum='TAMAMLANDI').count()
    devam_eden_ezber = EzberKaydi.objects.filter(durum='DEVAM').count()
    toplam_ezber = tamamlanan_ezber + devam_eden_ezber
    ezber_tamamlama_orani = round((tamamlanan_ezber / toplam_ezber * 100), 1) if toplam_ezber > 0 else 0
    
    # Seviye dağılımı
    seviye_dagilimi = {
        'HAZ1': Ogrenci.objects.filter(seviye='HAZ1').count(),
        'HAZ2': Ogrenci.objects.filter(seviye='HAZ2').count(),
        'HAZ3': Ogrenci.objects.filter(seviye='HAZ3').count(),
        'TEMEL': Ogrenci.objects.filter(seviye='TEMEL').count(),
        'ILERI': Ogrenci.objects.filter(seviye='ILERI').count(),
        'HAFIZLIK': Ogrenci.objects.filter(seviye='HAFIZLIK').count(),
    }
    
    # En başarılı 5 öğrenci
    en_basarili_5_ogrenci = []
    tum_ogrenciler = Ogrenci.objects.all()
    
    for ogrenci in tum_ogrenciler:
        ogrenci.ders_ortalamasi = SinavSonucu.objects.filter(
            ogrenci=ogrenci
        ).aggregate(ortalama=Avg('puan'))['ortalama'] or 0
    
    en_basarili_5_ogrenci = sorted(
        tum_ogrenciler, 
        key=lambda x: x.ders_ortalamasi, 
        reverse=True
    )[:5]
    
    # Son eklenen 5 öğrenci
    son_ogrenciler = Ogrenci.objects.all().order_by('-kayit_tarihi')[:5]
    
    # Son 5 yazı
    son_yazilar = yazi.objects.all().order_by('-id')[:5]
    
    context = {
        'toplam_yazi': toplam_yazi,
        'toplam_ogrenci': toplam_ogrenci,
        'sinif_ortalamasi': sinif_ortalamasi,
        'toplam_ezber': toplam_ezber,
        'tamamlanan_ezber': tamamlanan_ezber,
        'devam_eden_ezber': devam_eden_ezber,
        'ezber_tamamlama_orani': ezber_tamamlama_orani,
        'seviye_dagilimi': seviye_dagilimi,
        'en_basarili_5_ogrenci': en_basarili_5_ogrenci,
        'son_ogrenciler': son_ogrenciler,
        'son_yazilar': son_yazilar,
    }
    
    return render(request, 'admin_dashboard.html', context)


@login_required(login_url='login')
def yaziyaz(request):
    kategoriler = category.objects.all()
    msg = ""
    if request.method == "POST":
        title = request.POST["title"]
        description = request.POST["description"]
        imageUrl = request.FILES.get('image')
        isActive = request.POST.get("isActive", False)

        # Checkbox değerini doğru şekilde işleme
        isActive = True if isActive == "on" else False

        if title == "":
            msg+="Şeyma başlık girmek zorunlu"
            return render(request, 'yaziyaz.html',{
                    'error':True,
                    'msg':msg,
                    'kategoriler':kategoriler,}
                  )
        elif len(title) < 5 :
            msg+="Şeyma sence de başlık çok kısa değil mi?"
            return render(request, 'yaziyaz.html',{
                    'error':True,
                    'msg':msg,
                    'kategoriler':kategoriler,}
                  )
        elif description == "":
            msg+="Sence yazı içeriği olmadan paylaşım olur mu Şeyma?"
            return render(request, 'yaziyaz.html',{
                    'error':True,
                    'msg':msg,
                    'kategoriler':kategoriler,}
                  )
        elif len(description)<50:
            msg+="Şeyma yazı çok mu kısa oldu ne"
            return render(request, 'yaziyaz.html',{
                    'error':True,
                    'msg':msg,
                    'kategoriler':kategoriler,}
                  )

        yazilar = yazi(title=title, description=description, imageUrl=imageUrl, isActive=isActive)
        yazilar.save()
        
        return redirect('/blog')
        
    return render(request, 'yaziyaz.html',{
                  'kategoriler':kategoriler,}
                  )

def login(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method=="POST":
        form = AuthenticationForm(request, data=request.POST)

        username = request.POST["username"]
        password = request.POST["password"]

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login_auth(request, user)
            messages.add_message(request,messages.SUCCESS,"Giriş Başarılı")
            return redirect('home')
        else:
            messages.add_message(request,messages.WARNING,"Kullanıcı ismi veya parola hatalı")
            return render(request, 'giris.html')
    return render(request, 'giris.html')

from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, redirect

from django.core.mail import send_mail
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.shortcuts import render, redirect

@login_required(login_url='login')
def change_password(request):
    if request.method == "POST":
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            
            # Mail gönderme işlemi
            try:
                subject = 'Parolanız Güncellendi'
                message = f'''
Şeyma parola değişti bi haber vereyim dedim.

İyi günler dileriz,
{settings.SITE_NAME} Ekibi
                '''
                from_email = settings.DEFAULT_FROM_EMAIL
                recipient_list = [settings.CONTACT_EMAIL]
                
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=from_email,
                    recipient_list=recipient_list,
                    fail_silently=False,
                )
                
                messages.success(request, 'Parolanız başarıyla güncellendi.')
            except Exception as e:
                # Mail gönderilemezse hata mesajı göstermeden devam et
                messages.success(request, 'Parolanız başarıyla güncellendi.')
                # Hata loglama yapılabilir
                print(f"Mail gönderim hatası: {str(e)}")
            
            return redirect("parola_guncelle")
        else:
            # Form hatalarını messages ile göster
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{error}")
    else:
        form = PasswordChangeForm(request.user)
    
    return render(request, 'parola_guncelle.html', {'form': form})

def user_logout(request):
    messages.add_message(request,messages.SUCCESS,"Çıkış Yapıldı")
    logout(request)
    return render(request,'index.html')


from mainproject.models import Ogrenci, EzberKaydi

@login_required(login_url='login')
def ogrenci_duzenle(request, id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    
    # Tüm dersleri ve ezber sürelerini al
    tum_dersler = Ders.objects.all()
    tum_ezberler = EzberSuresi.objects.all().order_by('sira')
    seviyeler = Ogrenci.SEVIYE_CHOICES
    
    # Öğrencinin mevcut sınav sonuçlarını al
    sinav_sonuclari = SinavSonucu.objects.filter(ogrenci=ogrenci)
    
    # Öğrencinin mevcut ezber kayıtlarını al
    ezber_kayitlari = EzberKaydi.objects.filter(ogrenci=ogrenci)
    ezber_sozlugu = {ezber.sure_id: ezber for ezber in ezber_kayitlari}
    
    # Her ders için sınav sayısını hesapla ve ders nesnesine ekle
    for ders in tum_dersler:
        ders_sinavlari = sinav_sonuclari.filter(ders=ders)
        ders.sinav_sayisi = ders_sinavlari.count()
        ders.sinav_listesi = []
        for i, sinav in enumerate(ders_sinavlari, 1):
            ders.sinav_listesi.append({
                'index': i,
                'puan': sinav.puan,
                'id': sinav.id
            })
    
    if request.method == 'POST':
        # Temel bilgileri güncelle
        ogrenci.ad_soyad = request.POST.get('ad_soyad', '').title()
        ogrenci.ozel_notlar = request.POST.get('ozel_notlar', '')
        
        # Profil fotoğrafı güncelleme
        if 'profil_foto' in request.FILES:
            ogrenci.profil_foto = request.FILES['profil_foto']
        
        ogrenci.save()
        
        # Mevcut sınav sonuçlarını sil ve yenilerini ekle
        sinav_sonuclari.delete()
        for ders in tum_dersler:
            for i in range(1, 4):  # 3 sınav için
                puan = request.POST.get(f'sinav_puan_{ders.id}_{i}')
                if puan and puan.strip():
                    SinavSonucu.objects.create(
                        ogrenci=ogrenci,
                        ders=ders,
                        puan=int(puan),
                        sinav_tipi='GENEL',
                        aciklama=f"{i}. sınav"
                    )
        
        # Ezber kayıtlarını güncelle (silip yeniden oluşturma)
        for ezber in tum_ezberler:
            # Her ezber için verileri al
            durum = request.POST.get(f'ezber_durum_{ezber.id}', 'BASLAMADI')
            ilerleme = request.POST.get(f'ezber_ilerleme_{ezber.id}', 0)
            baslama_tarihi = request.POST.get(f'ezber_baslama_{ezber.id}') or None
            bitis_tarihi = request.POST.get(f'ezber_bitis_{ezber.id}') or None
            yorum = request.POST.get(f'ezber_yorum_{ezber.id}', '')
            
            # Mevcut kaydı kontrol et
            ezber_kaydi = ezber_sozlugu.get(ezber.id)
            
            if ezber_kaydi:
                # Kayıt varsa güncelle
                ezber_kaydi.durum = durum
                ezber_kaydi.ilerleme = ilerleme  # İlerleme alanını güncelle
                ezber_kaydi.baslama_tarihi = baslama_tarihi
                ezber_kaydi.bitis_tarihi = bitis_tarihi
                ezber_kaydi.yorum = yorum
                ezber_kaydi.save()
            else:
                # Kayıt yoksa yeni oluştur (sadece değerler varsayılandan farklıysa)
                if durum != 'BASLAMADI' or int(ilerleme) > 0 or baslama_tarihi or bitis_tarihi or yorum.strip():
                    EzberKaydi.objects.create(
                        ogrenci=ogrenci,
                        sure=ezber,
                        durum=durum,
                        ilerleme=ilerleme,  # İlerleme alanını ekle
                        baslama_tarihi=baslama_tarihi,
                        bitis_tarihi=bitis_tarihi,
                        yorum=yorum
                    )
        
        # Seviye güncellemesi yap
        update_ogrenci_seviye(ogrenci)
        
        messages.success(request, 'Öğrenci bilgileri ve tüm veriler güncellendi')
        return redirect('ogrenci_detay', id=ogrenci.id)
    
    # Ezber listesini hazırla
    ezber_listesi = []
    for ezber in tum_ezberler:
        ezber_kaydi = ezber_sozlugu.get(ezber.id)
        ezber_listesi.append({
            'id': ezber.id,
            'ad': ezber.ad,
            'sira': ezber.sira,
            'durum': ezber_kaydi.durum if ezber_kaydi else 'BASLAMADI',
            'baslama_tarihi': ezber_kaydi.baslama_tarihi if ezber_kaydi else None,
            'bitis_tarihi': ezber_kaydi.bitis_tarihi if ezber_kaydi else None,
            'yorum': ezber_kaydi.yorum if ezber_kaydi else '',
            'ilerleme': ezber_kaydi.ilerleme if ezber_kaydi else 0  # İlerleme alanını ekle
        })
    
    context = {
        'ogrenci': ogrenci,
        'tum_dersler': tum_dersler,
        'tum_ezberler': tum_ezberler,
        'seviyeler': seviyeler,
        'ezber_listesi': ezber_listesi,
    }
    
    return render(request, 'ogrenci_duzenle.html', context)

@login_required(login_url='login')
def ders_notu_ekle(request, id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    
    if request.method == 'POST':
        ders_id = request.POST.get('ders')
        not_degeri = request.POST.get('not_degeri')
        yorum = request.POST.get('yorum', '')
        tarih = request.POST.get('tarih', timezone.now().date())
        
        try:
            ders = Ders.objects.get(id=ders_id)
            not_degeri = int(not_degeri)
            
            if not 1 <= not_degeri <= 100:
                messages.error(request, 'Not değeri 1-100 arasında olmalıdır')
                return redirect('ogrenci_detay', id=id)
            
            # Aynı tarih ve ders için not var mı kontrol et
            mevcut_not = DersNotu.objects.filter(
                ogrenci=ogrenci, 
                ders=ders, 
                tarih=tarih
            ).first()
            
            if mevcut_not:
                mevcut_not.not_degeri = not_degeri
                mevcut_not.yorum = yorum
                mevcut_not.save()
                messages.success(request, f'{ders.get_tur_display()} notu güncellendi')
            else:
                DersNotu.objects.create(
                    ogrenci=ogrenci,
                    ders=ders,
                    not_degeri=not_degeri,
                    yorum=yorum,
                    tarih=tarih
                )
                messages.success(request, f'{ders.get_tur_display()} notu eklendi')
                
        except (Ders.DoesNotExist, ValueError):
            messages.error(request, 'Geçersiz veri girdiniz')
    
    return redirect('ogrenci_detay', id=id)

@login_required(login_url='login')
def ezber_ekle(request, id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    
    if request.method == 'POST':
        sure_id = request.POST.get('sure')
        baslama_tarihi = request.POST.get('baslama_tarihi', timezone.now().date())
        gunluk_ezber = request.POST.get('gunluk_ezber', 1)
        zorluk = request.POST.get('zorluk', 2)
        yorum = request.POST.get('yorum', '')
        
        try:
            sure = EzberSuresi.objects.get(id=sure_id)
            gunluk_ezber = int(gunluk_ezber)
            zorluk = int(zorluk)
            
            # Aktif ezberi kontrol et
            aktif_ezber = EzberKaydi.objects.filter(
                ogrenci=ogrenci, 
                tamamlandi=False
            ).first()
            
            if aktif_ezber:
                messages.warning(request, 'Öğrencinin zaten aktif bir ezberi var')
                return redirect('ogrenci_detay', id=id)
            
            EzberKaydi.objects.create(
                ogrenci=ogrenci,
                sure=sure,
                baslama_tarihi=baslama_tarihi,
                gunluk_ezber_miktari=gunluk_ezber,
                zorluk_seviyesi=zorluk,
                ogretmen_yorumu=yorum
            )
            
            messages.success(request, f'{sure.ad} ezberi başlatıldı')
            
        except (EzberSuresi.DoesNotExist, ValueError):
            messages.error(request, 'Geçersiz veri girdiniz')
    
    return redirect('ogrenci_detay', id=id)

@login_required(login_url='login')
def ezber_tamamla(request, id, ezber_id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    ezber_kaydi = get_object_or_404(EzberKaydi, id=ezber_id, ogrenci=ogrenci)
    
    if request.method == 'POST':
        ezber_kaydi.tamamlandi = True
        ezber_kaydi.bitis_tarihi = timezone.now().date()
        ezber_kaydi.save()
        
        messages.success(request, f'{ezber_kaydi.sure.ad} ezberi tamamlandı olarak işaretlendi')
    
    return redirect('ogrenci_detay', id=id)

@login_required(login_url='login')
def sinav_sonucu_ekle(request, id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    
    if request.method == 'POST':
        ders_id = request.POST.get('ders')
        puan = request.POST.get('puan')
        sinav_tipi = request.POST.get('sinav_tipi', 'QUIZ')
        detaylar = request.POST.get('detaylar', '')
        tarih = request.POST.get('tarih', timezone.now().date())
        
        try:
            ders = Ders.objects.get(id=ders_id)
            puan = int(puan)
            
            if not 1 <= puan <= 100:
                messages.error(request, 'Puan 1-100 arasında olmalıdır')
                return redirect('ogrenci_detay', id=id)
            
            SinavSonucu.objects.create(
                ogrenci=ogrenci,
                ders=ders,
                puan=puan,
                sinav_tipi=sinav_tipi,
                detaylar=detaylar,
                tarih=tarih
            )
            
            messages.success(request, f'{ders.get_tur_display()} sınav sonucu eklendi')
            
        except (Ders.DoesNotExist, ValueError):
            messages.error(request, 'Geçersiz veri girdiniz')
    
    return redirect('ogrenci_detay', id=id)

from django.db.models import Avg, Count, Sum, Q, F, ExpressionWrapper, DurationField
from django.utils import timezone
from datetime import timedelta
import json

def gemini_ogrenci_analizi(veri):
    """
    Gemini API'yi kullanarak öğrenci analizi yapar - Geliştirilmiş ve detaylı versiyon
    """
    # Önbellek anahtarı oluştur
    cache_key = f"ogrenci_analiz_{hash(str(veri))}"
    cached_response = cache.get(cache_key)
    
    if cached_response:
        return cached_response
    
    # Gemini API isteği
    url = "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent"
    
    headers = {
        'Content-Type': 'application/json',
        'X-goog-api-key': settings.GEMINI_API_KEY
    }
    
    # Geliştirilmiş ve detaylı prompt
    prompt = f"""
    Şeyma adında bir Kuran öğretmeni ve hafız olarak öğrenci analizi yapmanı istiyorum. 
    Amine Hatun Kuran Kursu'nda Hafızlık Hazırlık Öğretmenisin.
    
    LÜTFEN DİKKAT: Analizlerinizde motive edici veya genel ifadeler kullanma. 
    Sadece verilere dayalı, gerçekçi ve objektif değerlendirmeler yap.
    
    ÖĞRENCİ VERİLERİ:
    
    Öğrenci Bilgileri: {veri['ogrenci_bilgileri']}
    
    Sınav Performansı:
    - Genel Ortalama: {veri['sinav_ortalamasi']}
    - Ders Bazlı Ortalamalar: {veri['ders_bazli_ortalama']}
    - Sınıf Ortalaması: {veri['sinif_ortalamasi']}
    - Sınıf Sıralaması: {veri['sinif_siralamasi']}
    
    Ezber Performansı:
    - Tamamlanan Ezber: {veri['ezber_istatistikleri']['tamamlanan']}
    - Devam Eden Ezber: {veri['ezber_istatistikleri']['devam_eden']}
    - Toplam Ezber: {veri['ezber_istatistikleri']['toplam']}
    - Ezber Tamamlama Oranı: {veri['ezber_tamamlama_orani']}%
    - Sınıf Ezber Ortalaması: {veri['sinif_ezber_ortalamasi']}%
    
    Ezber Süre Analizi:
    {veri['ezber_sure_analizi']}
    
    Detaylı Ezber Süreleri:
    {veri['detayli_ezber_sureleri']}
    
    Kursa Katılım Süresi ve İlerleme Hızı:
    - Kursa başlama tarihi: {veri['katilma_tarihi']}
    - Toplam kursta geçen süre: {veri['toplam_kurs_suresi']}
    - Günlük ortalama ezber performansı: {veri['gunluk_ezber_ortalamasi']} sayfa/gün
    - Sınıfın günlük ortalama ezber performansı: {veri['sinif_gunluk_ortalamasi']} sayfa/gün
    
    Lütfen aşağıdaki analizleri YALNIZCA verilere dayalı olarak yap:
    
    1. ÖĞRENCİNİN MEVCUT DURUMU:
    - Kursa katılım süresine göre beklenen performans ile gerçekleşen performansı karşılaştır
    - Sınav performansını ders bazlı detaylı analiz et
    - Sınıf içindeki konumunu değerlendir
    - Güçlü ve zayıf yönlerini somut verilerle açıkla
    
    2. ZAMAN BAZLI PERFORMANS ANALİZİ:
    - Kursa katılım süresini diğer öğrencilerle karşılaştırarak normal seviyede olup olmadığını tespit et
    - Her bir sureyi ne kadar sürede tamamladığını detaylı analiz et
    - Ezber hızının zaman içindeki değişimini incele (hızlanma veya yavaşlama)
    - Günlük ezber ortalamasını sınıf ortalamasıyla karşılaştır
    
    3. EZBER PERFORMANS DEĞERLENDİRMESİ:
    - Ezber kalıcılığını ve tekrar düzenini analiz et
    - Hangi surelerde zorlandığını, hangilerinde daha başarılı olduğunu tespit et
    - Ezber verimliliğini (süre/sayfa oranı) hesapla ve değerlendir
    
    4. SINIF İÇİ KARŞILAŞTIRMA:
    - Sınav sonuçlarında sınıf içindeki yeri
    - Ezber performansında sınıf içindeki yeri
    - Katılım süresine göre beklenen performans ile gerçek performansın uyumluluğu
    - Genel performans sıralaması
    
    5. HAFIZLIK POTANSİYELİ ANALİZİ:
    - Mevcut ezber hızı ve kalitesine göre hafızlık süresi tahmini
    - Hafızlık için gereken minimum ve maksimum süre aralığı
    - Potansiyel riskler ve engeller
    - Verilen verilere dayalı olarak kişinin hafız olup olamayacağını değerlendir
    
    6. ÖĞRETMENE ÖZEL ÖNERİLER:
    - Bu öğrenciye özel çalışma stratejileri
    - Zayıf olduğu alanlara yönelik özel çözümler
    - Takip edilmesi gereken metrikler
    - Öğrencinin öğrenme hızına uygun beklenti yönetimi
    
    ANALİZ FORMATI:
    - Öğrencinin kursa katılım tarihi: {veri['katilma_tarihi']}
    - Toplam kursta geçen süre: {veri['toplam_kurs_suresi']}
    - Mevcut ezber durumu: {veri['ezber_istatistikleri']['tamamlanan']} sayfa
    - Günlük ortalama ezber: {veri['gunluk_ezber_ortalamasi']} sayfa/gün
    - Sınıf ortalaması: {veri['sinif_gunluk_ortalamasi']} sayfa/gün
    
    Bu verilere dayanarak öğrencinin:
    1. Katılım süresine göre performansını normal/üstün/zayıf olarak değerlendir
    2. Ezber hızının yeterliliğini ve sürdürülebilirliğini analiz et
    3. Sınıf içindeki konumunu katılım süresi perspektifinden yorumla
    4. Gerçekçi bir hafızlık tamamlama tahmini yap
    5. Öğrenme verimliliğini (süre/performans oranı) değerlendir
    
    Yanıtınız SADECE gerçekçi, veriye dayalı ve analitik olsun. 
    Motive edici ifadeler, genel geçer tavsiyeler veya klişeler KULLANMA.
    """
    
    payload = {
        "contents": [
            {
                "parts": [
                    {
                        "text": prompt
                    }
                ]
            }
        ],
        "generationConfig": {
            "temperature": 0.3,
            "topK": 20,
            "topP": 0.8,
            "maxOutputTokens": 4096,
        }
    }
    
    try:
        response = requests.post(url, headers=headers, json=payload, timeout=45)
        response.raise_for_status()
        
        result = response.json()
        
        # Yanıtı çıkar
        if (result.get('candidates') and 
            len(result['candidates']) > 0 and 
            result['candidates'][0].get('content') and
            result['candidates'][0]['content'].get('parts') and
            len(result['candidates'][0]['content']['parts']) > 0):
            
            cevap = result['candidates'][0]['content']['parts'][0]['text']
            
            # Metni formatla
            formatted_cevap = format_gemini_response(cevap)
            
            # Önbelleğe al (6 saat)
            cache.set(cache_key, formatted_cevap, 21600)
            
            return formatted_cevap
        else:
            return "**🤖 Analiz Hatası**\n\nÖğrenci analizi şu anda yapılamıyor. Lütfen daha sonra tekrar deneyin."
            
    except requests.exceptions.HTTPError as e:
        error_msg = f"HTTP hatası: {str(e)}"
        return f"**🤖 Teknik Sorun**\n\nAnaliz sırasında bir hata oluştu: {error_msg}"
        
    except requests.exceptions.Timeout:
        return "**⏰ Zaman Aşımı**\n\nAnaliz için zaman aşımı oluştu. Lütfen daha sonra tekrar deneyin."
        
    except requests.exceptions.ConnectionError:
        return "**🌐 Bağlantı Hatası**\n\nİnternet bağlantısı gerekiyor. Lütfen bağlantınızı kontrol edin."
        
    except Exception as e:
        return f"**❌ Beklenmeyen Hata**\n\nBir sorun oluştu: {str(e)}"


@login_required(login_url='login')
def ogrenci_detay(request, id):
    # Öğrenciyi ve ilişkili verileri tek sorguda al
    ogrenci = get_object_or_404(
        Ogrenci.objects.prefetch_related(
            Prefetch('sinavsonucu_set', queryset=SinavSonucu.objects.select_related('ders')),
            Prefetch('ezberkaydi_set', queryset=EzberKaydi.objects.select_related('sure'))
        ),
        id=id
    )
    
    # Sınav sonuçları ve ezber kayıtları artık önceden yüklenmiş olacak
    sinav_sonuclari = ogrenci.sinavsonucu_set.all()
    ezber_kayitlari = ogrenci.ezberkaydi_set.all()
    
    # Sınav ortalaması
    sinav_ortalamasi = sinav_sonuclari.aggregate(ortalama=Avg('puan'))['ortalama'] or 0
    
    # Ezber istatistikleri - daha verimli hesaplama
    ezber_durumlari = ezber_kayitlari.aggregate(
        tamamlanan=Count('id', filter=Q(durum='TAMAMLANDI')),
        devam_eden=Count('id', filter=Q(durum='DEVAM')),
        baslamayan=Count('id', filter=Q(durum='BASLAMADI'))
    )
    
    tamamlanan_ezberler = ezber_durumlari['tamamlanan'] or 0
    devam_eden_ezberler = ezber_durumlari['devam_eden'] or 0
    baslamayan_ezberler = ezber_durumlari['baslamayan'] or 0
    
    # Toplam ezber sayısını EzberSuresi'nden al
    toplam_ezber = EzberSuresi.objects.count()
    
    # Ezber süre analizi (sadece tamamlananlar için)
    tamamlanan_ezberler_list = ezber_kayitlari.filter(durum='TAMAMLANDI', baslama_tarihi__isnull=False, bitis_tarihi__isnull=False)
    ezber_sureleri = []
    detayli_ezber_sureleri = []
    
    for ezber in tamamlanan_ezberler_list:
        gun_sayisi = (ezber.bitis_tarihi - ezber.baslama_tarihi).days
        ezber_sureleri.append(gun_sayisi)
        detayli_ezber_sureleri.append({
            'sure_adi': ezber.sure.ad,
            'gun_sayisi': gun_sayisi,
            'tahmini_sure': ezber.sure.tahmini_sure,
            'fark': gun_sayisi - ezber.sure.tahmini_sure
        })
    
    ortalama_ezber_suresi = sum(ezber_sureleri) / len(ezber_sureleri) if ezber_sureleri else 0
    
    # Ders bazlı sınav ortalamaları
    ders_bazli_ortalama = {}
    for sinav in sinav_sonuclari:
        ders_adi = sinav.ders.get_tur_display()
        if ders_adi not in ders_bazli_ortalama:
            ders_bazli_ortalama[ders_adi] = []
        ders_bazli_ortalama[ders_adi].append(sinav.puan)
    
    for ders, puanlar in ders_bazli_ortalama.items():
        ders_bazli_ortalama[ders] = sum(puanlar) / len(puanlar)
    
    # Sınıf istatistikleri
    sinif_ortalamasi = SinavSonucu.objects.aggregate(ortalama=Avg('puan'))['ortalama'] or 0
    
    # Sınıf ezber ortalaması
    sinif_ezber_ortalamasi = EzberKaydi.objects.filter(durum='TAMAMLANDI').values('ogrenci').annotate(
        tamamlanan=Count('id')
    ).aggregate(ortalama=Avg('tamamlanan'))['ortalama'] or 0
    
    # Öğrencinin sınıf sıralaması - daha verimli hesaplama
    ogrenci_siralamasi_list = list(Ogrenci.objects.annotate(
        ortalama=Avg('sinavsonucu__puan'),
        tamamlanan_ezber=Count('ezberkaydi', filter=Q(ezberkaydi__durum='TAMAMLANDI'))
    ).order_by('-ortalama', '-tamamlanan_ezber').values('id', 'ad_soyad', 'ortalama', 'tamamlanan_ezber'))
    
    sinif_siralamasi = next((i+1 for i, o in enumerate(ogrenci_siralamasi_list) if o['id'] == ogrenci.id), 0)
    toplam_ogrenci_sayisi = len(ogrenci_siralamasi_list)
    
    # Tüm öğrenci verileri (grafikler için)
    tum_ogrenci_verileri = []
    for o in ogrenci_siralamasi_list:
        tum_ogrenci_verileri.append({
            'id': o['id'],
            'ad_soyad': o['ad_soyad'],
            'ortalama': o['ortalama'] or 0,
            'tamamlanan_ezber': o['tamamlanan_ezber'],
            'mevcut_ogrenci': o['id'] == ogrenci.id
        })
    
    # AI analizi için veri hazırla - sadece istek gelirse yap
    ai_analizi = None
    show_ai_analysis = request.GET.get('ai_analiz') == '1'
    
    if show_ai_analysis:
        katilma_tarihi = ogrenci.kayit_tarihi.strftime('%d/%m/%Y')
        toplam_kurs_suresi = (timezone.now().date() - ogrenci.kayit_tarihi).days
        
        # Günlük ortalama ezber performansı
        gunluk_ezber_ortalamasi = tamamlanan_ezberler / toplam_kurs_suresi if toplam_kurs_suresi > 0 else 0
        
        # Sınıfın günlük ortalama ezber performansı
        sinif_gunluk_ortalamasi = 0
        if toplam_ogrenci_sayisi > 0:
            tum_ogrenciler_ezber = Ogrenci.objects.annotate(
                tamamlanan_ezber=Count('ezberkaydi', filter=Q(ezberkaydi__durum='TAMAMLANDI'))
            )
            toplam_ezber_count = sum([o.ezberkaydi_set.filter(durum='TAMAMLANDI').count() for o in tum_ogrenciler_ezber])
            toplam_gun = sum([(timezone.now().date() - o.kayit_tarihi).days for o in tum_ogrenciler_ezber])
            sinif_gunluk_ortalamasi = toplam_ezber_count / toplam_gun if toplam_gun > 0 else 0
        
        ai_analiz_verisi = {
            'sinav_ortalamasi': sinav_ortalamasi,
            'ders_bazli_ortalama': ders_bazli_ortalama,
            'ezber_istatistikleri': {
                'tamamlanan': tamamlanan_ezberler,
                'devam_eden': devam_eden_ezberler,
                'toplam': toplam_ezber,
                'baslamayan': baslamayan_ezberler,
            },
            'ezber_tamamlama_orani': (tamamlanan_ezberler / toplam_ezber * 100) if toplam_ezber > 0 else 0,
            'ezber_sure_analizi': {
                'ortalama_ezber_suresi': ortalama_ezber_suresi,
                'en_hizli_ezber': min(ezber_sureleri) if ezber_sureleri else 0,
                'en_yavas_ezber': max(ezber_sureleri) if ezber_sureleri else 0,
                'toplam_ezber_suresi': sum(ezber_sureleri) if ezber_sureleri else 0,
            },
            'detayli_ezber_sureleri': detayli_ezber_sureleri,
            'sinif_ortalamasi': sinif_ortalamasi,
            'sinif_ezber_ortalamasi': sinif_ezber_ortalamasi,
            'sinif_siralamasi': f"{sinif_siralamasi}/{toplam_ogrenci_sayisi}",
            'ogrenci_bilgileri': {
                'ad_soyad': ogrenci.ad_soyad,
                'seviye': ogrenci.get_seviye_display(),
                'kayit_tarihi': katilma_tarihi,
                'kayit_suresi_gun': toplam_kurs_suresi,
                'ozel_notlar': ogrenci.ozel_notlar,
            },
            'katilma_tarihi': katilma_tarihi,
            'toplam_kurs_suresi': toplam_kurs_suresi,
            'gunluk_ezber_ortalamasi': round(gunluk_ezber_ortalamasi, 2),
            'sinif_gunluk_ortalamasi': round(sinif_gunluk_ortalamasi, 2)
        }
        
        # Gemini AI analizi
        ai_analizi = gemini_ogrenci_analizi(ai_analiz_verisi)
    
    context = {
        'ogrenci': ogrenci,
        'sinav_sonuclari': sinav_sonuclari,
        'ezber_kayitlari': ezber_kayitlari,
        'ai_analizi': mark_safe(ai_analizi) if ai_analizi else None,
        'show_ai_analysis': show_ai_analysis,
        'tamamlanan_ezberler': tamamlanan_ezberler,
        'devam_eden_ezberler': devam_eden_ezberler,
        'toplam_ezber': toplam_ezber,
        'baslamayan_ezberler': baslamayan_ezberler,
        'sinav_ortalamasi': sinav_ortalamasi,
        'ders_bazli_ortalama': ders_bazli_ortalama,
        'tum_ogrenci_verileri': json.dumps(tum_ogrenci_verileri),
        'sinif_ortalamasi': sinif_ortalamasi,
        'sinif_ezber_ortalamasi': sinif_ezber_ortalamasi,
        'sinif_siralamasi': sinif_siralamasi,
        'toplam_ogrenci_sayisi': toplam_ogrenci_sayisi,
        'ortalama_ezber_suresi': ortalama_ezber_suresi,
        'kayit_suresi_gun': (timezone.now().date() - ogrenci.kayit_tarihi).days,
        'simdi': timezone.now().date(),
    }
    
    return render(request, 'ogrenci_detay.html', context)

@login_required(login_url='login')
def ogrenci_listesi(request):
    # Tüm öğrencileri al, kayıt tarihine göre ters sırala (en yeni en üstte)
    tum_ogrenciler = Ogrenci.objects.all().order_by('-kayit_tarihi')
    
    # Arama fonksiyonelliği
    arama_terimi = request.GET.get('q')
    if arama_terimi:
        tum_ogrenciler = tum_ogrenciler.filter(ad_soyad__icontains=arama_terimi)
    
    # Seviye filtreleme
    seviye_filtre = request.GET.get('seviye')
    if seviye_filtre:
        tum_ogrenciler = tum_ogrenciler.filter(seviye=seviye_filtre)
    
    # Her öğrenci için istatistikleri hesapla
    ogrenci_ortalama_listesi = []
    for ogrenci in tum_ogrenciler:
        # Sınav ortalamasını hesapla (SinavSonucu modelini kullan)
        ogrenci.ders_ortalamasi = SinavSonucu.objects.filter(ogrenci=ogrenci).aggregate(
            Avg('puan')
        )['puan__avg'] or 0
        
        # Ezber sayısını hesapla
        ogrenci.tamamlanan_ezber_sayisi = EzberKaydi.objects.filter(
            ogrenci=ogrenci, durum='TAMAMLANDI'
        ).count()
        
        # Ortalama listesine ekle
        ogrenci_ortalama_listesi.append((ogrenci, ogrenci.ders_ortalamasi))
    
    # Sınıf ortalamasını hesapla (tüm öğrencilerin sınav ortalamalarının ortalaması)
    sinif_ortalamasi = 0
    if ogrenci_ortalama_listesi:
        toplam_ortalama = sum(ortalama for _, ortalama in ogrenci_ortalama_listesi)
        sinif_ortalamasi = toplam_ortalama / len(ogrenci_ortalama_listesi)
    
    # En başarılı öğrenciyi bul (sınav ortalaması en yüksek olan)
    en_basarili_ogrenci = None
    if ogrenci_ortalama_listesi:
        en_basarili_ogrenci, en_yuksek_ortalama = max(ogrenci_ortalama_listesi, key=lambda x: x[1])
        en_basarili_ogrenci.ders_ortalamasi = en_yuksek_ortalama
    
    # En başarılı 5 öğrenciyi bul
    en_basarili_5_ogrenci = []
    if ogrenci_ortalama_listesi:
        # Ortalamaya göre sırala (yüksekten düşüğe)
        ogrenci_ortalama_listesi.sort(key=lambda x: x[1], reverse=True)
        # İlk 5 öğrenciyi al
        en_basarili_5_ogrenci = ogrenci_ortalama_listesi[:5]
    
    # Toplam tamamlanan ezber sayısı
    toplam_tamamlanan_ezber = EzberKaydi.objects.filter(durum='TAMAMLANDI').count()
    
    # Seviye dağılımını hesapla
    seviye_dagilimi = {
        'HAZ1': 0, 'HAZ2': 0, 'HAZ3': 0, 
        'TEMEL': 0, 'ILERI': 0, 'HAFIZLIK': 0
    }
    for ogrenci in tum_ogrenciler:
        if ogrenci.seviye in seviye_dagilimi:
            seviye_dagilimi[ogrenci.seviye] += 1
    
    # Ders bazlı ortalamaları hesapla
    ders_ortalamalari = {}
    dersler = Ders.objects.all()
    for ders in dersler:
        ortalama = SinavSonucu.objects.filter(ders=ders).aggregate(Avg('puan'))['puan__avg'] or 0
        ders_ortalamalari[ders.get_tur_display()] = ortalama
    
    # Sayfalama
    sayfa = request.GET.get('sayfa', 1)
    paginator = Paginator(tum_ogrenciler, 10)
    
    try:
        ogrenciler = paginator.page(sayfa)
    except PageNotAnInteger:
        ogrenciler = paginator.page(1)
    except EmptyPage:
        ogrenciler = paginator.page(paginator.num_pages)
    
    # Seviye seçeneklerini al
    seviyeler = Ogrenci.SEVIYE_CHOICES
    
    return render(request, 'ogrenci_listesi.html', {
        'ogrenciler': ogrenciler,
        'seviyeler': seviyeler,
        'arama_terimi': arama_terimi,
        'seviye_filtre': seviye_filtre,
        'sinif_ortalamasi': sinif_ortalamasi,
        'en_basarili_ogrenci': en_basarili_ogrenci,
        'en_basarili_5_ogrenci': en_basarili_5_ogrenci,
        'toplam_tamamlanan_ezber': toplam_tamamlanan_ezber,
        'seviye_dagilimi': seviye_dagilimi,
        'ders_ortalamalari': ders_ortalamalari,
    })

@login_required(login_url='login')
def ogrenci_not_ekle(request,id):
    ogrenci = get_object_or_404(Ogrenci, id=id)
    if request.method == 'POST':
        ogrenci.ozel_notlar = request.POST.get('ozel_not', '')
        ogrenci.save()
        messages.success(request, 'Notlar başarıyla kaydedildi')
    return redirect('ogrenci_detay', id=ogrenci.id)

@login_required(login_url='login')
def ogrenci_ekle(request):
    # Tüm gerekli verileri al
    tum_dersler = Ders.objects.all()
    tum_ezberler = EzberSuresi.objects.all().order_by('sira')
    
    if request.method == 'POST':
        # Öğrenciyi oluştur
        yeni_ogrenci = Ogrenci()
        yeni_ogrenci.ad_soyad = request.POST.get('ad_soyad', '').title()
        yeni_ogrenci.ozel_notlar = request.POST.get('ozel_notlar', '')
        
        # Seviye otomatik olarak HAZ1'den başlayacak
        yeni_ogrenci.seviye = 'HAZ1'
        
        # Profil fotoğrafı güncelleme
        if 'profil_foto' in request.FILES:
            yeni_ogrenci.profil_foto = request.FILES['profil_foto']
        
        yeni_ogrenci.save()
        
        # Sınav puanlarını güncelle - her ders için
        for ders in tum_dersler:
            # Ana sınav puanı
            puan = request.POST.get(f'sinav_puan_{ders.id}_1')
            if puan and puan.strip():
                SinavSonucu.objects.create(
                    ogrenci=yeni_ogrenci,
                    ders=ders,
                    puan=int(puan),
                    sinav_tipi='GENEL',
                    aciklama="Ana sınav"
                )
            
            # Ek sınav puanları
            for i in range(2, 4):
                puan = request.POST.get(f'sinav_puan_{ders.id}_{i}')
                if puan and puan.strip():
                    SinavSonucu.objects.create(
                        ogrenci=yeni_ogrenci,
                        ders=ders,
                        puan=int(puan),
                        sinav_tipi='GENEL',
                        aciklama=f"{i}. ek sınav"
                    )
        
        # Ezber kayıtlarını güncelle
        for ezber in tum_ezberler:
            durum = request.POST.get(f'ezber_durum_{ezber.id}')
            ilerleme = request.POST.get(f'ezber_ilerleme_{ezber.id}', 0)
            
            if durum and durum != 'BASLAMADI':
                baslama_tarihi = request.POST.get(f'ezber_baslama_{ezber.id}')
                bitis_tarihi = request.POST.get(f'ezber_bitis_{ezber.id}')
                yorum = request.POST.get(f'ezber_yorum_{ezber.id}')
                
                # Tarih alanları boş ise None yap
                baslama_tarihi = baslama_tarihi if baslama_tarihi else None
                bitis_tarihi = bitis_tarihi if bitis_tarihi else None
                
                EzberKaydi.objects.create(
                    ogrenci=yeni_ogrenci,
                    sure=ezber,
                    durum=durum,
                    ilerleme=ilerleme,  # İlerleme alanını ekliyoruz
                    baslama_tarihi=baslama_tarihi,
                    bitis_tarihi=bitis_tarihi,
                    yorum=yorum
                )
            elif ilerleme and int(ilerleme) > 0:
                # Sadece ilerleme varsa kayıt oluştur
                baslama_tarihi = request.POST.get(f'ezber_baslama_{ezber.id}')
                bitis_tarihi = request.POST.get(f'ezber_bitis_{ezber.id}')
                yorum = request.POST.get(f'ezber_yorum_{ezber.id}')
                
                baslama_tarihi = baslama_tarihi if baslama_tarihi else None
                bitis_tarihi = bitis_tarihi if bitis_tarihi else None
                
                EzberKaydi.objects.create(
                    ogrenci=yeni_ogrenci,
                    sure=ezber,
                    durum='DEVAM',  # İlerleme varsa otomatik olarak DEVAM yap
                    ilerleme=ilerleme,
                    baslama_tarihi=baslama_tarihi,
                    bitis_tarihi=bitis_tarihi,
                    yorum=yorum
                )
        
        # Seviye güncellemesi yap
        update_ogrenci_seviye(yeni_ogrenci)
        
        messages.success(request, 'Öğrenci başarıyla eklendi')
        return redirect('ogrenci_listesi')
    
    # GET isteği için
    context = {
        'tum_dersler': tum_dersler,
        'tum_ezberler': tum_ezberler,
    }
    return render(request, 'yeni_ogrenci.html', context)

def update_ogrenci_seviye(ogrenci):
    """Öğrenci seviyesini ezber durumuna göre otomatik günceller"""
    tamamlanan_ezberler = EzberKaydi.objects.filter(
        ogrenci=ogrenci, 
        durum='TAMAMLANDI'
    ).count()
    
    # Seviye belirleme kriterleri
    if tamamlanan_ezberler >= 10:
        ogrenci.seviye = 'ILERI'
    elif tamamlanan_ezberler >= 7:
        ogrenci.seviye = 'TEMEL'
    elif tamamlanan_ezberler >= 4:
        ogrenci.seviye = 'HAZ3'
    elif tamamlanan_ezberler >= 2:
        ogrenci.seviye = 'HAZ2'
    else:
        ogrenci.seviye = 'HAZ1'
    
    ogrenci.save()

@login_required
def ogrenci_sil(request, ogrenci_id):
    ogrenci = get_object_or_404(Ogrenci, id=ogrenci_id)
    
    if request.method == 'POST':
        ogrenci.delete()
        messages.success(request, f'{ogrenci.ad_soyad} başarıyla silindi')
        return redirect('ogrenci_listesi')
    
    return render(request, 'ogrenci_sil_onay.html', {'ogrenci': ogrenci})
